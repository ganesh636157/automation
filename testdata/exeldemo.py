import openpyxl
book=openpyxl.load_workbook("C:\\Users\\Windows\\PycharmProjects\\pythondemo1.xlsx")
sheet=book.active
#cell=sheet.cell(row=1,column=2)
#print(cell.value)
#sheet.cell(row=2,column=2).value="Ganesh"
#print(sheet.cell(row=2,column=2).value)
#print(sheet['A3'].value)
Country1=[]
#Country12=[]
for i in range(2,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        Country1.append((sheet.cell(row=i,column=j).value))

print(Country1)


