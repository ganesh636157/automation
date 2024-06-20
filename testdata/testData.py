import openpyxl

class test:
    #test_data=[("India","Aruba","Fiji"),("Aruba","India","Fiji")] #or
    book = openpyxl.load_workbook("C:\\Users\\Windows\\PycharmProjects\\pythondemo1.xlsx")
    sheet = book.active
    Country1 = []
    Country2 = []
    test_data=[]
    for i in range(2, sheet.max_row+1):
        for j in range(2,3):
            Country1.append((sheet.cell(row=i, column=j).value))
    #print(Country1)
    for i in range(2, sheet.max_row + 1):
        for j in range(3, sheet.max_column + 1):
            Country2.append((sheet.cell(row=i, column=j).value))
    #print(Country2)
    test_data=[Country1,Country2]
    print(test_data)

